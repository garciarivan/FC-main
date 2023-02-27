from PyQt5 import QtCore, QtGui, QtWidgets
import sys
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import column_index_from_string
import pandas as pd
from PyQt5.QtWidgets import QTableWidgetItem
from openpyxl import load_workbook
import lib_gui
from lib_gui import pop_up
from docx import Document
from docx.shared import Pt
from docx.shared import Inches
from docx.shared import RGBColor
from docx.shared import Cm
import os
import datetime
from mailmerge import MailMerge
import shutil

template = r'C:\Users\medico.RSD\Documents\FC-main\playground\template.docx'
filename = 'assets\listado_pacientes.xlsx'
ruta_carmen_ball = r'C:\Users\medico.RSD\Documents\FC-main\playground\CARMEN BALLESTEROS'
ruta_cris_plas = r'C:\Users\medico.RSD\Documents\FC-main\playground\CRISTINA PLASENCIA'
ruta_fab_rosby = r'C:\Users\medico.RSD\Documents\FC-main\playground\FABIOLA ROSBY'
ruta_jose_a_fuentes = r'C:\Users\medico.RSD\Documents\FC-main\playground\JA FUENTES'
ruta_peromingo = r'C:\Users\medico.RSD\Documents\FC-main\playground\JA PEROMINGO'
ruta_leoncio = r'C:\Users\medico.RSD\Documents\FC-main\playground\LEONCIO BERNAL'
ruta_angeles_perez = r'C:\Users\medico.RSD\Documents\FC-main\playground\M ANGELES PEREZ'
ruta_remesal = r'C:\Users\medico.RSD\Documents\FC-main\playground\MARIA REMESAL'
ruta_natalia = r'C:\Users\medico.RSD\Documents\FC-main\playground\NATALIA CLEMENTE'
ruta_rosario = r'C:\Users\medico.RSD\Documents\FC-main\playground\ROSARIO GONZALEZ'
ruta_verdion = r'C:\Users\medico.RSD\Documents\FC-main\playground\SUSANA VERDION'
rutas = [ruta_natalia, ruta_leoncio, ruta_remesal, ruta_cris_plas, ruta_jose_a_fuentes, ruta_verdion, ruta_angeles_perez, ruta_carmen_ball, ruta_peromingo, ruta_rosario, ruta_fab_rosby]
MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

lista_map = ['Natalia Clemente', 'Leoncio Bernal', 'Maria Remesal', 'Cristina Plasencia', 'JA Fuentes', 'Susana Verdion',
             'M Angeles Perez', \
             'Carmen Ballesteros', 'JA Peromingo', 'Rosario Gonzalez', 'Fabiola Rosby']


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(771, 560)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.listView = QtWidgets.QTableWidget(self.centralwidget)
        self.listView.setGeometry(QtCore.QRect(30, 110, 362, 382))
        self.listView.setObjectName("listView")

        self.listView.setColumnCount(3)
        self.listView.setRowCount(self.extract_cells())
        item = QtWidgets.QTableWidgetItem('NOMBRE')
        self.listView.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem('FN')
        self.listView.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem('MAP')
        self.listView.setHorizontalHeaderItem(2, item)

        self.load_residents()
        

        self.label_name = QtWidgets.QLabel(self.centralwidget)
        self.label_name.setGeometry(QtCore.QRect(441, 286, 201, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_name.setFont(font)
        self.label_name.setObjectName("label_name")
        self.label_name.setVisible(False)

        self.textEdit_name = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit_name.setGeometry(QtCore.QRect(441, 312, 251, 31))
        self.textEdit_name.setObjectName("textEdit_name")
        self.textEdit_name.setVisible(False)

        self.label_fn = QtWidgets.QLabel(self.centralwidget)
        self.label_fn.setGeometry(QtCore.QRect(441, 352, 251, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_fn.setFont(font)
        self.label_fn.setObjectName("label_fn")
        self.label_fn.setVisible(False)

        self.textEdit_fn = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit_fn.setGeometry(QtCore.QRect(441, 378, 251, 31))
        self.textEdit_fn.setObjectName("textEdit_fn")
        self.textEdit_fn.setVisible(False)

        self.label_map = QtWidgets.QLabel(self.centralwidget)
        self.label_map.setGeometry(QtCore.QRect(441, 418, 251, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_map.setFont(font)
        self.label_map.setObjectName("label_map")
        self.label_map.setVisible(False)

        self.combobox_map = QtWidgets.QComboBox(self.centralwidget)
        self.combobox_map.setGeometry(QtCore.QRect(441, 444, 251, 31))
        self.combobox_map.setObjectName("textEdit_map")
        self.combobox_map.setVisible(False)
        for map in lista_map:
            self.combobox_map.addItem(map)

        self.textEdit = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit.setGeometry(QtCore.QRect(30, 70, 251, 31))
        self.textEdit.setObjectName("textEdit")

        self.textEdit_2 = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit_2.setGeometry(QtCore.QRect(441, 110, 311, 81))
        self.textEdit_2.setObjectName("textEdit_2")

        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(441, 80, 201, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")

        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(30, 40, 131, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")

        # boton añadir
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(30, 500, 121, 41))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.add_resident)
        
        # boton confirmar
        self.pushButton_add = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_add.setGeometry(QtCore.QRect(441, 500, 121, 41))
        self.pushButton_add.setObjectName("pushButton_add")
        self.pushButton_add.setVisible(False)
        self.pushButton_add.clicked.connect(self.confirm)
        
        # boton eliminar
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(160, 500, 121, 41))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.clicked.connect(self.remove_resident)
        
        # boton solicitar
        self.pushButton_3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_3.setGeometry(QtCore.QRect(632, 200, 121, 41))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_3.clicked.connect(self.request)
        
        # boton resetear semana
        self.pushButton_4 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_4.setGeometry(QtCore.QRect(441, 200, 121, 41))
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_4.clicked.connect(self.reset_week)
        

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 771, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        
    def extract_cells(self):
        self.spreadsheet = load_workbook(filename)
        self.sheet = self.spreadsheet.active
        rows = self.sheet.max_row
        self.spreadsheet.save(filename)
        return rows

    
    def load_residents(self):
        self.spreadsheet = load_workbook(filename)
        self.sheet = self.spreadsheet.active
        for row in self.sheet:
            if all(cell.value is None for cell in row):
                self.sheet.delete_rows(row[0].row, 1)
        self.sheet.calculate_dimension()
        self.spreadsheet.save(filename)
        rows = self.sheet.max_row - 1
        self.listView.setRowCount(rows)
        self.listView.clearContents()
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=2, max_col=3)):
            if str(row[0].value) == 'None':
                continue
            self.listView.setItem(row_num, 0, QTableWidgetItem(str(row[0].value)))
            self.listView.setItem(row_num, 1, QTableWidgetItem(self.format_date_of_birth(str(row[1].value))))
            self.listView.setItem(row_num, 2, QTableWidgetItem(str(row[2].value)))


        
        

    def add_resident(self):
        self.label_name.setVisible(True)
        self.textEdit_name.setVisible(True)
        self.label_fn.setVisible(True)
        self.textEdit_fn.setVisible(True)
        self.label_map.setVisible(True)
        self.combobox_map.setVisible(True)
        self.pushButton_add.setVisible(True)
     

    def confirm(self):
        
        self.sheet['A' + str(self.sheet.max_row + 1)] = self.textEdit_name.toPlainText()
        self.sheet['B' + str(self.sheet.max_row)] = self.textEdit_fn.toPlainText()
        self.sheet['C' + str(self.sheet.max_row)] = self.combobox_map.currentText()
        self.textEdit_name.setText('')
        self.textEdit_fn.setText('') 
        # Ordenar las filas por orden alfabético en la primera columna
        sort_worksheet(self.sheet, 1)
        self.spreadsheet.save(filename)
        pop_up.information(self, 'BIEN!!', 'El residente de ha añadido con éxito')
        self.load_residents()

    def format_date_of_birth(self, date):
        self.format1 = date.split(' ')[0]
        self.format2 = self.format1.split('-')
        return '/'.join(self.format2[::-1])

    def remove_resident(self):
        selected_items = self.listView.selectedItems()
        if not selected_items:
            pop_up.error(self, 'ATENCIÓN', 'No se ha seleccionado ningún residente.')
            return
        # Obtener el índice de la fila seleccionada en la tabla
        row_index_view = self.listView.currentRow()
        # Ajustar el índice de la fila para obtener el número de fila correspondiente en la hoja de cálculo
        row_index_sheet = row_index_view + 2
        # Eliminar la fila correspondiente en la hoja de Excel
        self.spreadsheet = load_workbook(filename)
        self.sheet = self.spreadsheet.active
        self.sheet.delete_rows(row_index_sheet, 1)
        self.spreadsheet.save(filename)
        # Actualizar la tabla
        self.load_residents()
        pop_up.information(self, 'BIEN!!', 'El residente de ha eliminado con éxito')



    def request(self):
        selected_items = self.listView.selectedItems()
        if not selected_items:
            pop_up.error(self, 'ATENCIÓN', 'No se ha seleccionado ningún residente.')
            return
        selected_item = self.listView.item(self.listView.currentRow(), 0).text()
        new_row = None
        # leer archivo Word acorde con MAP correspondiente
        names_to_paths = {
            'Natalia Clemente': ruta_natalia,
            'Leoncio Bernal': ruta_leoncio,
            'Maria Remesal': ruta_remesal,
            'Cristina Plasencia': ruta_cris_plas,
            'JA Fuentes': ruta_jose_a_fuentes,
            'Susana Verdion': ruta_verdion,
            'M Angeles Perez': ruta_angeles_perez,
            'Carmen Ballesteros': ruta_carmen_ball,
            'JA Peromingo': ruta_peromingo,
            'Rosario Gonzalez': ruta_rosario,
            'Fabiola Rosby': ruta_fab_rosby,
        }

        name = self.listView.item(self.listView.currentRow(), 2).text()
        if name in names_to_paths:
            ruta = os.path.join(names_to_paths[name], name.upper() + ' ' + fechado_archivo(datetime.datetime.now()) + '.docx')
            document = Document(ruta)
        else:
            pop_up.error(self, 'ATENCION', 'Debe verificar ortografia de MAP')

        # agregar una nueva fila a la tabla del documento
        table = document.tables[0]
        if table.table.cell(-1, 0).text.strip():
            new_row = table.add_row()
        else:
            new_row = table.rows[-1]
        new_row.cells[0].text = str(selected_item + '\n' + self.listView.item(self.listView.currentRow(), 1).text()).upper()
        new_row.cells[1].text = self.textEdit_2.toPlainText().upper()
        new_row.cells[2].text = 'DEMANDA'
        pop_up.information(self, 'Aviso', 'Solicitud guardada correctamente.')
        # guardar el archivo Word
        document.save(ruta)

        

    def reset_week(self):
        documento = MailMerge(template)
        fecha_actual = datetime.datetime.now()  
        dias_a_lunes = fecha_actual.weekday()
        dias_a_domingo = 6 - dias_a_lunes  
        fecha_lunes = fecha_actual - datetime.timedelta(days=dias_a_lunes)
        fecha_domingo = fecha_actual + datetime.timedelta(days=dias_a_domingo)
        fecha_lunes_anterior = fecha_lunes - datetime.timedelta(days=7)
        fecha_domingo_anterior = fecha_domingo - datetime.timedelta(days=7)
        for x, y in zip(lista_map, rutas):
            documento.merge(map=x.upper(), lunes=str(fecha_lunes.day), domingo=str(fecha_domingo.day), mes=MESES[fecha_domingo.month -1].upper(), year=str(fecha_domingo.year))
            documento.write(y + '\\' + x.upper() + ' ' + fechado_archivo(datetime.datetime.now()) + '.docx')
            shutil.move(y + '\\' + x.upper() + ' ' + fechado_archivo(datetime.date(fecha_lunes_anterior.year, fecha_lunes_anterior.month, fecha_lunes_anterior.day)) + '.docx', y + '\\' + 'LISTADOS ANTIGUOS' + '\\' + x.upper() + ' ' + fechado_archivo(datetime.date(fecha_lunes_anterior.year, fecha_lunes_anterior.month, fecha_lunes_anterior.day)) + '.docx')
        pop_up.information(self, 'HECHO!', 'Semana reseteada')

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "SOLICITUDES A MAP"))
        self.textEdit.setHtml(_translate("MainWindow",
                                         "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
                                         "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
                                         "p, li { white-space: pre-wrap; }\n"
                                         "</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:8.25pt; font-weight:400; font-style:normal;\">\n"
                                         "<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">Buscar</p></body></html>"))
        self.label.setText(_translate("MainWindow", "Escriba la solicitud a continuacion:"))
        self.label_2.setText(_translate("MainWindow", "Listado de residentes:"))
        self.label_name.setText(_translate("MainWindow", "Nombre y Apellidos:"))
        self.label_fn.setText(_translate("MainWindow", "Fecha de Nacimiento:"))
        self.label_map.setText(_translate("MainWindow", "MAP:"))
        self.pushButton.setText(_translate("MainWindow", "Añadir"))
        self.pushButton_2.setText(_translate("MainWindow", "Eliminar"))
        self.pushButton_3.setText(_translate("MainWindow", "SOLICITAR"))
        self.pushButton_4.setText(_translate("MainWindow", "RESETEAR\n"
                                                           "SEMANA"))
        self.pushButton_add.setText(_translate("MainWindow", "CONFIRMAR"))


def main():
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())



def sort_worksheet(worksheet, col):
    data = worksheet.values
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)
    df = df.sort_values(cols[int(col)-1])
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)


def fechado_archivo(fecha):
      
    dias_a_lunes = fecha.weekday()
    dias_a_domingo = 6 - dias_a_lunes  
    fecha_lunes = fecha - datetime.timedelta(days=dias_a_lunes)
    fecha_domingo = fecha + datetime.timedelta(days=dias_a_domingo)
    return f'LISTADO RECETAS SEMANAL DEL {fecha_lunes.day} AL {fecha_domingo.day} DE {MESES[fecha_domingo.month -1].upper()} {fecha.year}'


if __name__ == "__main__":
    main()
    
